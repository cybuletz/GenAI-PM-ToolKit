import os
import base64

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
TEXT_EXTENSIONS = {".txt", ".md", ".py", ".ts", ".js", ".json", ".yaml", ".yml", ".xml", ".csv"}

MIME_MAP = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


class FileProcessor:
    """
    Processes uploaded files into a format suitable for the GitHub Models API.
    Images are base64-encoded. Text files are read as plain text.
    PDF, DOCX, XLSX are extracted to plain text.
    """

    def process(self, file_path: str) -> dict:
        ext = os.path.splitext(file_path)[1].lower()

        if ext in IMAGE_EXTENSIONS:
            return self._process_image(file_path, ext)
        elif ext in TEXT_EXTENSIONS:
            return self._process_text(file_path)
        elif ext == ".pdf":
            return self._process_pdf(file_path)
        elif ext == ".docx":
            return self._process_docx(file_path)
        elif ext == ".xlsx":
            return self._process_xlsx(file_path)
        else:
            return {"type": "text", "data": f"[Unsupported file type: {ext}]"}

    def _process_image(self, path: str, ext: str) -> dict:
        with open(path, "rb") as f:
            data = base64.b64encode(f.read()).decode("utf-8")
        return {"type": "image", "mime": MIME_MAP.get(ext, "image/png"), "data": data}

    def _process_text(self, path: str) -> dict:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return {"type": "text", "data": f.read()}

    def _process_pdf(self, path: str) -> dict:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(path)
            text = "\n".join(page.get_text() for page in doc)
            return {"type": "text", "data": text}
        except ImportError:
            return {"type": "text", "data": "[PDF support requires PyMuPDF: pip install PyMuPDF]"}

    def _process_docx(self, path: str) -> dict:
        try:
            from docx import Document
            doc = Document(path)
            text = "\n".join(p.text for p in doc.paragraphs)
            return {"type": "text", "data": text}
        except ImportError:
            return {"type": "text", "data": "[DOCX support requires python-docx: pip install python-docx]"}

    def _process_xlsx(self, path: str) -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join(str(c) if c is not None else "" for c in row))
            return {"type": "text", "data": "\n".join(lines)}
        except ImportError:
            return {"type": "text", "data": "[XLSX support requires openpyxl: pip install openpyxl]"}
